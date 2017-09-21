'''
Created on Sep 18, 2017

@author: kishlay.mishra
'''
from openpyxl import load_workbook

class Excel_Data:
    
    def get_data(self,val):
        wb = load_workbook(filename = 'input.xlsx')
        ws=wb['Data']
        ws.sheet_properties.tabColor = "1D0E98"
        res=''
        if val=='UserName':
            res=ws['A2'].value
        elif val=='Password':
            res=ws['B2'].value
        elif val=='Emailid':
            res=ws['C2'].value
        elif val=='FirstName':
            res=ws['D2'].value
        elif val=='LastName':
            res=ws['E2'].value
        elif val=='Address':
            res=ws['F2'].value
        elif val=='City':
            res=ws['G2'].value
        elif val=='Country':
            res=ws['H2'].value
        elif val=='State':
            res=ws['I2'].value
        elif val=='Pincode':
            res=ws['J2'].value
        elif val=='PhoneNumber':
            res=ws['K2'].value
        
        wb.close()
        
        return res
    
    def write_result(self,val):
        wbR = load_workbook(filename = 'input.xlsx')
        dest_filename = 'input.xlsx'
        
        sheetR=wbR.active
        
        if('Result' in wbR.sheetnames):
            sheetR=wbR['Result']
        else:
            sheetR=wbR.create_sheet(title='Result')
        
        sheetR.sheet_properties.tabColor = "d6ca9e"
            
        sheetR['A1']=val
        wbR.save(filename = dest_filename)
        